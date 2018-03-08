import openpyxl

from common import Level, translation_cols
from phrase import Phrase
from view import View

"""
Wrapper class for the excel reader engine - a kind of abstraction/separation layer for modularity
"""


class Excel:
    def __init__(self, path):
        self.path = path
        self.engine = openpyxl.load_workbook(path)

    def get_sheet_by_name(self, name):
        return self.engine[name]

    def get_sheet_names(self):
        return self.engine.sheetnames

    def get_value(self, current_sheet, col_row):
        return current_sheet[col_row].value

    def fill_color(self, current_sheet, col_row, color):
        current_sheet[col_row].fill = color

    def save_changes(self, vocabulary):

        View.print_("Please wait while saving...")
        for sheet_name in self.get_sheet_names():
            for level in [Level.LOW, Level.MID,Level.HIGH, Level.UNKNOWN]:
                for key, phrase in vocabulary[sheet_name][level].iteritems():
                    self.fill_color(
                        color=Level.to_style[phrase.get_level()],
                        current_sheet=self.get_sheet_by_name(sheet_name),
                        col_row=phrase.get_position()
                    )
        self.engine.save(self.path)

    def get_wb_as_dict(self):

        # TODO - memoization?
        wb_dict = {}

        for sheet_name in self.get_sheet_names():

            wb_dict[sheet_name] = self._init_sheet_dict()

            # current_sheet = vocbulary_wb.get_sheet_by_name(sheet_name)
            current_sheet = self.get_sheet_by_name(sheet_name)
            # get number of rows
            rows = list(range(3, current_sheet.max_row + 1))
            for row in rows:
                try:
                    # get phrase
                    starting_position = 'A' + str(row)
                    current_phrase = self.get_value(current_sheet, starting_position)
                    current_context = self.get_value(current_sheet, 'L' + str(row))
                    current_translation = [current_sheet[col + str(row)].value for col in translation_cols if
                                           current_sheet[col + str(row)].value is not None]
                    level = Level.get_phrase_level(current_sheet, starting_position)
                    phrase = Phrase(value=current_phrase,
                                    context=current_context,
                                    translation=current_translation,
                                    position=starting_position, level=level)
                    wb_dict[sheet_name][level][phrase.get_value()] = phrase
                    # wb_dict[current_sheet][level][phrase.get_value()] = phrase.to_dict()

                except Exception:
                    View.print_("Error reading row in {} sheet".format(sheet_name))
                    pass

        return wb_dict

        # iterate each sheet
        # each row
        # create a phrase from each row
        pass

    @staticmethod
    def _init_sheet_dict():
        return {
                Level.HIGH: {},
                Level.MID: {},
                Level.LOW: {},
                Level.UNKNOWN: {}
        }
