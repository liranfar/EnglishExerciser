import openpyxl

my_red = openpyxl.styles.colors.Color(rgb='FFD33D3D')
red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
my_green = openpyxl.styles.colors.Color(rgb='FF55996F')
green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
my_yellow = openpyxl.styles.colors.Color(rgb='FFD0D637')
yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
my_blank = openpyxl.styles.colors.Color(rgb='FFFFFFFF')
blank_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_blank)